import xlrd 
import time
import openpyxl
from openpyxl.styles import Font, Color, Alignment, Border, Side, colors
from openpyxl.styles import NamedStyle

def SampleForm(ws,listDevice_OneEvent,sheet):
	ws.cell(1,1).value = "Event"
	ws.cell(1,2).value = "Detected on"
	ws.cell(1,3).value = "Group"
	ws.cell(1,4).value = "Device"
	ws.cell(1,5).value = "Task"
	ws.cell(1,6).value = "Description"
	ws.cell(1,7).value = "IP address"

	big_red_text = Font(bold=True,color=colors.RED, size=10)
	sheet["A1"].font = big_red_text
	sheet["B1"].font = big_red_text
	sheet["C1"].font = big_red_text
	sheet["D1"].font = big_red_text
	sheet["E1"].font = big_red_text
	sheet["F1"].font = big_red_text
	sheet["G1"].font = big_red_text

	for x in range(0,len(listDevice_OneEvent)):
		ws.cell(x+2,1).value = listDevice_OneEvent[x][0]
		ws.cell(x+2,2).value = listDevice_OneEvent[x][1]
		ws.cell(x+2,3).value = listDevice_OneEvent[x][2]
		ws.cell(x+2,4).value = listDevice_OneEvent[x][3]
		ws.cell(x+2,5).value = listDevice_OneEvent[x][4]
		ws.cell(x+2,6).value = listDevice_OneEvent[x][5]
		ws.cell(x+2,7).value = listDevice_OneEvent[x][6]

def SampleFormOne(ws,listDevice_OneEvent,sheet):
	ws.cell(1,1).value = "Event"
	ws.cell(1,2).value = "Detected on"
	ws.cell(1,3).value = "Group"
	ws.cell(1,4).value = "Device"
	ws.cell(1,5).value = "Task"
	ws.cell(1,6).value = "Description"
	ws.cell(1,7).value = "IP address"
	ws.cell(1,8).value = "IP address Attack"

	big_red_text = Font(bold=True,color=colors.RED, size=10)
	sheet["A1"].font = big_red_text
	sheet["B1"].font = big_red_text
	sheet["C1"].font = big_red_text
	sheet["D1"].font = big_red_text
	sheet["E1"].font = big_red_text
	sheet["F1"].font = big_red_text
	sheet["G1"].font = big_red_text
	sheet["H1"].font = big_red_text

	for x in range(0,len(listDevice_OneEvent)):
		ws.cell(x+2,1).value = listDevice_OneEvent[x][0]
		ws.cell(x+2,2).value = listDevice_OneEvent[x][1]
		ws.cell(x+2,3).value = listDevice_OneEvent[x][2]
		ws.cell(x+2,4).value = listDevice_OneEvent[x][3]
		ws.cell(x+2,5).value = listDevice_OneEvent[x][4]
		ws.cell(x+2,6).value = listDevice_OneEvent[x][5]
		ws.cell(x+2,7).value = listDevice_OneEvent[x][6]
		ws.cell(x+2,8).value = getIpAttack(listDevice_OneEvent[x][5])

def getIpAttack(stringDes):
	i_start = stringDes.find('TCP from ')
	i_end   = stringDes.find("to ")
	if i_start == -1 or i_end == -1:
		return "Several different sources"
	else:
		result =  stringDes[i_start+9:i_end]
		return result

def Create_Sheet_New(filename,eventList,deviceList):		
	wb = openpyxl.load_workbook(filename)
	for index in range(0,len(eventList)):
		nameSheet = eventList[index]
		if len(nameSheet) > 31:
			nameSheet = nameSheet[0:31]
		wb.create_sheet(nameSheet) # Sheet Name
		ws = wb.worksheets[index+2] 		   # Sheet Index
		listDevice_OneEvent = deviceList[index]
		# if eventList[index] == "Network attack detected" or eventList[index] == "Disinfection impossible" or eventList[index] == "Malicious object detected" :
		if eventList[index] == "Network attack detected":
			SampleFormOne(ws,listDevice_OneEvent,ws)
		else:
			SampleForm(ws,listDevice_OneEvent,ws)
	wb.save(filename)

def EventCount(filename, eventList):
	loc = (filename) 
	wb = xlrd.open_workbook(loc)
	sheet = wb.sheet_by_index(1) 
	listEventName = []
	for i in range(0,len(eventList)):
		tempList = []
		if eventList[i] == "Network attack detected" or eventList[i] == "Disinfection impossible" or eventList[i] == "Malicious object detected" :
			DeviceCountSpecial(sheet,eventList[i],tempList)
		else:
			DeviceCount(sheet,eventList[i],tempList)
		listEventName.append(tempList)
	return listEventName

def DeviceCountSpecial(sheet,eventListName,tempList):
	for r in range(1,sheet.nrows):
		s_value = sheet.cell_value(r,2)
		if s_value == eventListName and sheet.cell_value(r,5) != "Managed devices":
			tempList.append([s_value,sheet.cell_value(r,3),sheet.cell_value(r,5),sheet.cell_value(r,6),sheet.cell_value(r,7),sheet.cell_value(r,8),sheet.cell_value(r,9),""])

def DeviceCount(sheet,eventListName,tempList):
	for r in range(1,sheet.nrows):
		s_value = sheet.cell_value(r,2)
		if s_value == eventListName and sheet.cell_value(r,5) != "Managed devices":
			tempList.append([s_value,sheet.cell_value(r,3),sheet.cell_value(r,5),sheet.cell_value(r,6),sheet.cell_value(r,7),sheet.cell_value(r,8),sheet.cell_value(r,9)])
		

def EventList(filename):
	loc = (filename) 
	wb = xlrd.open_workbook(loc)
	sheet = wb.sheet_by_index(1) 
	# print(len(wb.sheet_names()))sheet.nrows
	# print(sheet.nrows)
	listEvent = []
	for r in range(1,sheet.nrows):
		if sheet.cell_value(r,5) != "Managed devices":
			s_value = sheet.cell_value(r,2)
			listEvent.append(s_value)
	listEvent = set(listEvent)
	listEvent = list(listEvent)
	listEvent.sort()
	# print(type(listEvent))
	# print(listEvent)
	# print(len(listEvent))
	return listEvent

def sortEventList(eventList):
	j = 0
	for i in range(0,len(eventList)):
		if eventList[i] == "Network attack detected":
			temp = eventList[j]
			eventList[j] = eventList[i]
			eventList[i] = temp
			j=j+1
	for i in range(0,len(eventList)):
		if eventList[i] == eventList[i] == "Malicious object detected":
			temp = eventList[j]
			eventList[j] = eventList[i]
			eventList[i] = temp
			j=j+1
	for i in range(0,len(eventList)):
		if eventList[i] == eventList[i] == "Disinfection impossible":
			temp = eventList[j]
			eventList[j] = eventList[i]
			eventList[i] = temp
			j=j+1
	return eventList

def function(filename):
	loc = (filename) 
	try:
		wb = xlrd.open_workbook(loc) 
	except Exception as err:
            raise ProcessingException('Invalid Excel file: %s' % err)
	# print(len(wb.sheet_names()))
	sheet = wb.sheet_by_index(1) 
	listDevice = []
	listGroup = []
	listError = []
	for r in range(1,sheet.nrows):
		listError.append((sheet.cell_value(r,0),sheet.cell_value(r,2))) 
		listDevice.append((sheet.cell_value(r,0),sheet.cell_value(r,2),sheet.cell_value(r,1))) 
		listGroup.append((sheet.cell_value(r,0),sheet.cell_value(r,2),sheet.cell_value(r,5))) 
	s_listError = set(listError)
	s_listDevice = set(listDevice)
	s_listGroup = set(listGroup)
	# print(len(list(s_listError)))
	s_listError = list(s_listError)
	s_listDevice = list(s_listDevice)
	s_listGroup = list(s_listGroup)
	# print(s_listError)
	c_listError = len(s_listError)
	c_listDevice = len(s_listDevice)
	c_listGroup = len(s_listGroup)
	listKey =[]
	for i in range(0,c_listError):
		l = list(s_listError[i])
		l.append(0)
		l.append(0)
		l.append(0)
		listKey.append(l)
	for r in range(0,c_listDevice):
		for i in range(0,c_listError):
			if s_listDevice[r][0] == listKey[i][0] and s_listDevice[r][1] == listKey[i][1]:
				listKey[i][3] = listKey[i][3] + 1
	for r in range(0,c_listGroup):
		for i in range(0,c_listError):
			if s_listGroup[r][0] == listKey[i][0] and s_listGroup[r][1] == listKey[i][1]:
				listKey[i][4] = listKey[i][4] + 1
	for r in range(1,sheet.nrows):
		for i in range(0,c_listError):
			if sheet.cell_value(r,0) == listKey[i][0] and sheet.cell_value(r,2) == listKey[i][1]:
				listKey[i][2] = listKey[i][2] + 1
	wb = openpyxl.load_workbook(filename)
	ws = wb.worksheets[0]
	ws.cell(7,7).value = ""
	ws.cell(7,8).value = ""
	for x in range(0,len(listKey)):
		ws.cell(x+8,1).value = listKey[x][0]
		ws.cell(x+8,2).value = listKey[x][2]
		ws.cell(x+8,3).value = listKey[x][1]
		ws.cell(x+8,4).value = listKey[x][3]
		ws.cell(x+8,5).value = listKey[x][2]
		ws.cell(x+8,6).value = listKey[x][4]
		ws.cell(x+8,7).value = ""
		ws.cell(x+8,8).value = ""
	wb.save(filename)
	# print("Success!!!")

def main(filename):
	eventList = sortEventList(EventList(filename))
	# print(len(eventList))
	print("1...")
	# print(eventList[0])
	deviceList = EventCount(filename,eventList)
	# print(deviceList)
	print("2...")
	# print(len(deviceList[0][0]))
	# print(deviceList[0][0][5])
	Create_Sheet_New(filename,eventList,deviceList)
	function(filename)
	print("3")


if __name__== "__main__":
	print("Library needs to be installed: xlrd, openpyxl.")
	print("Enter filename (no write .xlsx) : (Only run filetype .xlsx)")
	# fileName=input() 
	fileName = "E:\\VNPT Security Inter\\ALL\\ToolCount\\12-19"
	fileName = fileName + ".xlsx"
	start_time = time.time()
	main(fileName)
	end_time = time.time()
	print("Time : ", (end_time - start_time), "s")